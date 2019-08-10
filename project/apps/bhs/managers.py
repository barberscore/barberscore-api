import time
from datetime import date

# Third-Party
from algoliasearch_django.decorators import disable_auto_indexing
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from phonenumber_field.validators import validate_international_phonenumber

# Django
from django.apps import apps
from django.db.models import Manager
from django.db.models import Q
from django.db.models import F
from django.db.models import Min
from django.db.models import Max
from django.db.models import When
from django.db.models import Subquery
from django.db.models import OuterRef
from django.db.models import CharField
from django.db.models import IntegerField
from django.db.models import DateField
from django.db.models import Case
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model

User = get_user_model()


class PersonManager(Manager):
    def update_or_create_from_human(self, human):
        # Extract
        if isinstance(human, dict):
            mc_pk = human['id']
            first_name = human['first_name']
            middle_name = human['middle_name']
            last_name = human['last_name']
            nick_name = human['nick_name']
            email = human['email']
            birth_date = human['birth_date']
            home_phone = human['home_phone']
            cell_phone = human['cell_phone']
            work_phone = human['work_phone']
            bhs_id = human['bhs_id']
            gender = human['gender']
            part = human['part']
            mon = human['mon']
            is_deceased = human['is_deceased']
            is_honorary = human['is_honorary']
            is_suspended = human['is_suspended']
            is_expelled = human['is_expelled']
        else:
            mc_pk = str(human.id)
            first_name = human.first_name
            middle_name = human.middle_name
            last_name = human.last_name
            nick_name = human.nick_name
            email = human.email
            birth_date = human.birth_date
            home_phone = human.home_phone
            cell_phone = human.cell_phone
            work_phone = human.work_phone
            bhs_id = human.bhs_id
            gender = human.gender
            part = human.part
            mon = human.mon
            is_deceased = human.is_deceased
            is_honorary = human.is_honorary
            is_suspended = human.is_suspended
            is_expelled = human.is_expelled

        # Transform
        inactive = any([
            is_deceased,
            is_honorary,
            is_suspended,
            is_expelled,
        ])
        if inactive:
            status = self.model.STATUS.inactive
        else:
            status = self.model.STATUS.active

        prefix = first_name.rpartition('Dr.')[1].strip()
        first_name = first_name.rpartition('Dr.')[2].strip()
        last_name = last_name.partition('II')[0].strip()
        suffix = last_name.partition('II')[1].strip()
        last_name = last_name.partition('III')[0].strip()
        suffix = last_name.partition('III')[1].strip()
        last_name = last_name.partition('DDS')[0].strip()
        suffix = last_name.partition('DDS')[1].strip()
        last_name = last_name.partition('Sr')[0].strip()
        suffix = last_name.partition('Sr')[1].strip()
        last_name = last_name.partition('Jr')[0].strip()
        suffix = last_name.partition('Jr')[1].strip()
        last_name = last_name.partition('M.D.')[0].strip()
        suffix = last_name.partition('M.D.')[1].strip()
        if nick_name == first_name:
            nick_name = ""

        try:
            validate_international_phonenumber(home_phone)
        except ValidationError:
            home_phone = ""
        try:
            validate_international_phonenumber(cell_phone)
        except ValidationError:
            cell_phone = ""
        try:
            validate_international_phonenumber(work_phone)
        except ValidationError:
            work_phone = ""

        if gender:
            gender = getattr(self.model.GENDER, gender, None)
        else:
            gender = None
        if part:
            part = getattr(self.model.PART, part, None)
        else:
            part = None

        try:
            validate_international_phonenumber(home_phone)
        except ValidationError:
            home_phone = ""

        try:
            validate_international_phonenumber(cell_phone)
        except ValidationError:
            cell_phone = ""

        try:
            validate_international_phonenumber(work_phone)
        except ValidationError:
            work_phone = ""

        is_deceased = bool(is_deceased)


        defaults = {
            'status': status,
            'prefix': prefix,
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name,
            'suffix': suffix,
            'nick_name': nick_name,
            'email': email,
            'birth_date': birth_date,
            'home_phone': home_phone,
            'cell_phone': cell_phone,
            'work_phone': work_phone,
            'bhs_id': bhs_id,
            'gender': gender,
            'part': part,
            'is_deceased': is_deceased,
            'mon': mon,
        }
        # Update or create
        person, created = self.update_or_create(
            mc_pk=mc_pk,
            defaults=defaults,
        )
        return person, created

    def delete_orphans(self, humans):
        # Delete Orphans
        orphans = self.filter(
            mc_pk__isnull=False,
        ).exclude(
            mc_pk__in=humans,
        )
        t = orphans.count()
        orphans.delete()
        return t

    def export_orphans(self, cursor=None):
        ps = self.filter(
            email__isnull=True,
            user__isnull=False,
        )
        if cursor:
            ps = ps.filter(
                modified__gte=cursor,
            )
        return ps

    def export_adoptions(self, cursor=None):
        ps = self.filter(
            email__isnull=False,
            user__isnull=True,
        )
        if cursor:
            ps = ps.filter(
                modified__gte=cursor,
            )
        return ps


class GroupManager(Manager):
    def update_or_create_from_structure(self, structure):
        # Extract
        if isinstance(structure, dict):
            mc_pk = structure['id']
            name = structure['name']
            kind = structure['kind']
            gender = structure['gender']
            division = structure['division']
            bhs_id = structure['bhs_id']
            legacy_code = structure['chapter_code']
            website = structure['website']
            email = structure['email']
            main_phone = structure['phone']
            fax_phone = structure['fax']
            facebook = structure['facebook']
            twitter = structure['twitter']
            youtube = structure['youtube']
            pinterest = structure['pinterest']
            flickr = structure['flickr']
            instagram = structure['instagram']
            soundcloud = structure['soundcloud']
            preferred_name = structure['preferred_name']
            visitor_information = structure['visitor_information']
            established_date = structure['established_date']
            status_id = structure['status_id']
            parent_pk = structure['parent_id']
        else:
            mc_pk = str(structure.id)
            name = structure.name
            kind = structure.kind
            gender = structure.gender
            division = structure.division
            bhs_id = structure.bhs_id
            legacy_code = structure.chapter_code
            website = structure.website
            email = structure.email
            main_phone = structure.phone
            fax_phone = structure.fax
            facebook = structure.facebook
            twitter = structure.twitter
            youtube = structure.youtube
            pinterest = structure.pinterest
            flickr = structure.flickr
            instagram = structure.instagram
            soundcloud = structure.soundcloud
            preferred_name = structure.preferred_name
            visitor_information = structure.visitor_information
            established_date = structure.established_date
            status_id = structure.status_id
            parent_pk = structure.parent_id


        # Transform
        status_map = {
            '64ad817f-f3c6-4b09-a1b0-4bd569b15d03': self.model.STATUS.inactive, # revoked
            'd9e3e257-9eca-4cbf-959f-149cca968349': self.model.STATUS.inactive, # suspended
            '6e3c5cc6-0734-4edf-8f51-40d3a865a94f': self.model.STATUS.inactive, # merged
            'bd4721e7-addd-4854-9888-8a705725f748': self.model.STATUS.inactive, # closed
            'e04744e6-b743-4247-92c2-2950855b3a93': self.model.STATUS.inactive, # expired
            '55a97973-02c3-414a-bbef-22181ad46e85': self.model.STATUS.active, # pending
            'bb1ee6f6-a2c5-4615-b6ad-76130c37b1e6': self.model.STATUS.active, # pending voluntary
            'd7102af8-013a-40e7-bc85-0b00766ed124': self.model.STATUS.active, # awaiting
            'f3facc00-1990-4c68-9052-39e066906a38': self.model.STATUS.active, # prospective
            '4bfee76f-3110-4c32-bade-e5044fdd5fa2': self.model.STATUS.active, # licensed
            '7b9e5e34-a7c5-4f1e-9fc5-656caa74b3c7': self.model.STATUS.active, # active
        }
        status = status_map.get(status_id, None)

        # Re-construct dangling article
        name = name.strip() if name else ""
        parsed = name.partition(", The")
        name = "The {0}".format(parsed[0]) if parsed[1] else parsed[0]

        preferred_name = "{0} (NAME APPROVAL PENDING)".format(preferred_name.strip()) if preferred_name else ''
        name = name if name else preferred_name

        if not name:
            name = "(UNKNOWN)"

        # AIC
        aic_map = {
            503061: "Signature",
            500983: "After Hours",
            501972: "Main Street",
            501329: "Forefront",
            500922: "Instant Classic",
            304772: "Musical Island Boys",
            500000: "Masterpiece",
            501150: "Ringmasters",
            317293: "Old School",
            286100: "Storm Front",
            500035: "Crossroads",
            297201: "OC Times",
            299233: "Max Q",
            302244: "Vocal Spectrum",
            299608: "Realtime",
            6158: "Gotcha!",
            2496: "Power Play",
            276016: "Four Voices",
            5619: "Michigan Jake",
            6738: "Platinum",
            3525: "FRED",
            5721: "Revival",
            2079: "Yesteryear",
            2163: "Nightlife",
            4745: "Marquis",
            3040: "Joker's Wild",
            1259: "Gas House Gang",
            2850: "Keepsake",
            1623: "The Ritz",
            3165: "Acoustix",
            1686: "Second Edition",
            492: "Chiefs of Staff",
            1596: "Interstate Rivals",
            1654: "Rural Route 4",
            406: "The New Tradition",
            1411: "Rapscallions",
            1727: "Side Street Ramblers",
            545: "Classic Collection",
            490: "Chicago News",
            329: "Boston Common",
            4034: "Grandma's Boys",
            318: "Bluegrass Student Union",
            362: "Most Happy Fellows",
            1590: "Innsiders",
            1440: "Happiness Emporium",
            1427: "Regents",
            627: "Dealer's Choice",
            1288: "Golden Staters",
            1275: "Gentlemen's Agreement",
            709: "Oriole Four",
            711: "Mark IV",
            2047: "Western Continentals",
            1110: "Four Statesmen",
            713: "Auto Towners",
            715: "Four Renegades",
            1729: "Sidewinders",
            718: "Town and Country 4",
            719: "Gala Lads",
            1871: "The Suntones",
            722: "Evans Quartet",
            724: "Four Pitchikers",
            726: "Gaynotes",
            729: "Lads of Enchantment",
            731: "Confederates",
            732: "Four Hearsemen",
            736: "The Orphans",
            739: "Vikings",
            743: "Four Teens",
            746: "Schmitt Brothers",
            748: "Buffalo Bills",
            750: "Mid-States Four",
            753: "Pittsburghers",
            756: "Doctors of Harmony",
            759: "Garden State Quartet",
            761: "Misfits",
            764: "Harmony Halls",
            766: "Four Harmonizers",
            770: "Elastic Four",
            773: "Chord Busters",
            775: "Flat Foot Four",
            776: "Bartlsesville Barflies",
        }

        # Overwrite status for AIC
        if bhs_id in aic_map:
            status = -5

        # Overwrite name for AIC
        name = aic_map.get(bhs_id, name)

        kind_map = {
            'quartet': self.model.KIND.quartet,
            'chorus': self.model.KIND.chorus,
            'chapter': self.model.KIND.chapter,
            'group': self.model.KIND.noncomp,
            'district': self.model.KIND.district,
            'organization': self.model.KIND.international,
        }
        kind = kind_map.get(kind, None)

        legacy_code = legacy_code if legacy_code else ""

        gender_map = {
            'men': self.model.GENDER.male,
            'women': self.model.GENDER.female,
            'mixed': self.model.GENDER.mixed,
        }
        gender = gender_map.get(gender, self.model.GENDER.male)

        division_map = {
            'EVG Division I': self.model.DIVISION.evgd1,
            'EVG Division II': self.model.DIVISION.evgd2,
            'EVG Division III': self.model.DIVISION.evgd3,
            'EVG Division IV': self.model.DIVISION.evgd4,
            'EVG Division V': self.model.DIVISION.evgd5,
            'FWD Arizona': self.model.DIVISION.fwdaz,
            'FWD Northeast': self.model.DIVISION.fwdne,
            'FWD Northwest': self.model.DIVISION.fwdnw,
            'FWD Southeast': self.model.DIVISION.fwdse,
            'FWD Southwest': self.model.DIVISION.fwdsw,
            'LOL 10000 Lakes': self.model.DIVISION.lol10l,
            'LOL Division One': self.model.DIVISION.lolone,
            'LOL Northern Plains': self.model.DIVISION.lolnp,
            'LOL Packerland': self.model.DIVISION.lolpkr,
            'LOL Southwest': self.model.DIVISION.lolsw,
            'MAD Central': self.model.DIVISION.madcen,
            'MAD Northern': self.model.DIVISION.madnth,
            'MAD Southern': self.model.DIVISION.madsth,
            'NED Granite and Pine': self.model.DIVISION.nedgp,
            'NED Mountain': self.model.DIVISION.nedmtn,
            'NED Patriot': self.model.DIVISION.nedpat,
            'NED Sunrise': self.model.DIVISION.nedsun,
            'NED Yankee': self.model.DIVISION.nedyke,
            'SWD Northeast': self.model.DIVISION.swdne,
            'SWD Northwest': self.model.DIVISION.swdnw,
            'SWD Southeast': self.model.DIVISION.swdse,
            'SWD Southwest': self.model.DIVISION.swdsw,
        }
        division = division_map.get(division, None)

        visitor_information = visitor_information.strip() if visitor_information else ''

        if parent_pk:
            parent = self.get(
                mc_pk=parent_pk,
            )
        else:
            parent = None

        if parent:
            if parent.kind == 'organization':
                representing_raw = legacy_code
            elif parent.kind == 'district':
                representing_raw = parent.legacy_code
            elif parent.kind == 'chapter':
                representing_raw = parent.parent.legacy_code
            else:
                representing_raw = None
        elif kind == 'organization':
            representing_raw = 'BHS'
        else:
            representing_raw = None

        representing_map = {
            'BHS': self.model.REPRESENTING.bhs,
            'CAR': self.model.REPRESENTING.car,
            'CSD': self.model.REPRESENTING.csd,
            'DIX': self.model.REPRESENTING.dix,
            'EVG': self.model.REPRESENTING.evg,
            'FWD': self.model.REPRESENTING.fwd,
            'ILL': self.model.REPRESENTING.ill,
            'JAD': self.model.REPRESENTING.jad,
            'LOL': self.model.REPRESENTING.lol,
            'MAD': self.model.REPRESENTING.mad,
            'NED': self.model.REPRESENTING.ned,
            'NSC': self.model.REPRESENTING.nsc,
            'ONT': self.model.REPRESENTING.ont,
            'PIO': self.model.REPRESENTING.pio,
            'RMD': self.model.REPRESENTING.rmd,
            'SLD': self.model.REPRESENTING.sld,
            'SUN': self.model.REPRESENTING.sun,
            'SWD': self.model.REPRESENTING.swd,
        }
        representing = representing_map.get(representing_raw, None)

        defaults = {
            'status': status,
            'name': name,
            'kind': kind,
            'gender': gender,
            'representing': representing,
            'division': division,
            'bhs_id': bhs_id,
            'code': legacy_code,
            'website': website,
            'email': email,
            'phone': main_phone,
            'fax_phone': fax_phone,
            'facebook': facebook,
            'twitter': twitter,
            'youtube': youtube,
            'pinterest': pinterest,
            'flickr': flickr,
            'instagram': instagram,
            'soundcloud': soundcloud,
            'visitor_information': visitor_information,
            'start_date': established_date,
            'parent': parent,
        }

        # Load
        group, created = self.update_or_create(
            mc_pk=mc_pk,
            defaults=defaults,
        )
        return group, created

    def delete_orphans(self, structures):
        # Delete Orphans
        orphans = self.filter(
            mc_pk__isnull=False,
        ).exclude(
            mc_pk__in=structures,
        )
        t = orphans.count()
        orphans.delete()
        return t

    def sort_tree(self):
        self.all().update(tree_sort=None)
        root = self.get(kind=self.model.KIND.international)
        i = 1
        root.tree_sort = i
        with disable_auto_indexing(model=self.model):
            root.save()
        for child in root.children.order_by('kind', 'code', 'name'):
            i += 1
            child.tree_sort = i
            with disable_auto_indexing(model=self.model):
                child.save()
        orgs = self.filter(
            kind__in=[
                self.model.KIND.chapter,
                self.model.KIND.chorus,
                self.model.KIND.quartet,
            ]
        ).order_by(
            'kind',
            'name',
        )
        for org in orgs:
            i += 1
            org.tree_sort = i
            with disable_auto_indexing(model=self.model):
                org.save()
        return

    def denormalize(self, cursor=None):
        groups = self.filter(status=self.model.STATUS.active)
        if cursor:
            groups = groups.filter(
                modified__gte=cursor,
            )
        for group in groups:
            group.denormalize()
            with disable_auto_indexing(model=self.model):
                group.save()
        return

    def update_seniors(self):
        quartets = self.filter(
            kind=self.model.KIND.quartet,
            status__gt=0,
            mc_pk__isnull=False,
        )

        for quartet in quartets:
            prior = quartet.is_senior
            is_senior = quartet.get_is_senior()
            if prior != is_senior:
                quartet.is_senior = is_senior
                with disable_auto_indexing(model=self.model):
                    quartet.save()
        return

    def get_quartets(self):
        wb = Workbook()
        ws = wb.active
        fieldnames = [
            'PK',
            'Name',
            'Kind',
            'Organization',
            'District',
            'Chapter(s)',
            'Senior?',
            'BHS ID',
            'Code',
            'Status',
        ]
        ws.append(fieldnames)
        groups = self.filter(
            status=self.model.STATUS.active,
            kind=self.model.KIND.quartet,
        ).order_by('name')
        for group in groups:
            pk = str(group.pk)
            name = group.name
            kind = group.get_kind_display()
            organization = "FIX"
            district = group.district
            chapters = group.chapters
            is_senior = group.is_senior
            is_youth = group.is_youth
            bhs_id = group.bhs_id
            code = group.code
            status = group.get_status_display()
            row = [
                pk,
                name,
                kind,
                organization,
                district,
                chapters,
                is_senior,
                is_youth,
                bhs_id,
                code,
                status,
            ]
            ws.append(row)
        file = save_virtual_workbook(wb)
        content = ContentFile(file)
        return content


class AwardManager(Manager):
    def sort_tree(self):
        self.all().update(tree_sort=None)
        awards = self.order_by(
            '-status',  # Actives first
            'district',  # Basic BHS Hierarchy
            '-kind', # Quartet, Chorus
            'gender', #Male, mixed
            F('age').asc(nulls_first=True), # Null, Senior, Youth
            'level', #Championship, qualifier
            'is_novice',
            'name', # alpha
        )
        i = 0
        for award in awards:
            i += 1
            award.tree_sort = i
            award.save()
        return

    def get_awards(self):
        wb = Workbook()
        ws = wb.active
        fieldnames = [
            'ID',
            'District',
            'Division',
            'Name',
            'Kind',
            'Gender',
            'Season',
            'Level',
            'Single',
            'Spots',
            'Threshold',
            'Minimum',
            'Advance',
        ]
        ws.append(fieldnames)
        awards = self.select_related(
            # 'group',
        ).filter(
            status__gt=0,
        ).order_by('tree_sort')
        for award in awards:
            pk = str(award.id)
            district = award.get_district_display()
            division = award.get_division_display()
            name = award.name
            kind = award.get_kind_display()
            gender = award.get_gender_display()
            season = award.get_season_display()
            level = award.get_level_display()
            single = award.is_single
            spots = award.spots
            threshold = award.threshold
            minimum = award.minimum
            advance = award.advance
            row = [
                pk,
                district,
                division,
                name,
                kind,
                gender,
                season,
                level,
                single,
                spots,
                threshold,
                minimum,
                advance,
            ]
            ws.append(row)
        file = save_virtual_workbook(wb)
        content = ContentFile(file)
        return content


class ChartManager(Manager):
    def get_report(self):
        wb = Workbook()
        ws = wb.active
        fieldnames = [
            'PK',
            'Title',
            'Arrangers',
            'Composers',
            'Lyricists',
            'Holders',
            'Status',
        ]
        ws.append(fieldnames)
        charts = self.order_by('title', 'arrangers')
        for chart in charts:
            pk = str(chart.pk)
            title = chart.title
            arrangers = chart.arrangers
            composers = chart.composers
            lyricists = chart.lyricists
            holders = chart.holders
            status = chart.get_status_display()
            row = [
                pk,
                title,
                arrangers,
                composers,
                lyricists,
                holders,
                status,
            ]
            ws.append(row)
        file = save_virtual_workbook(wb)
        content = ContentFile(file)
        return content

